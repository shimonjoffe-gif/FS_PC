import openpyxl
import sys

path = r"Копия ФС Lite и ПК для предварительной оценки заказного проекта.xlsx"
wb = openpyxl.load_workbook(path, data_only=False)

for name in wb.sheetnames:
    if "ПРОФ" in name and "Очередь 1" in name:
        print("Sheet:", name)
        ws = wb[name]
        for row in range(51, 58):
            c = ws[f"C{row}"]
            b = ws[f"B{row}"]
            print(f"Row {row}: B={b.value!r} C={c.value!r}")
        break

ws2 = wb["1.Содержание проекта"]
print("\n--- Contract criteria rows 173-182 ---")
for row in range(173, 183):
    b = ws2[f"B{row}"]
    c = ws2[f"C{row}"]
    d = ws2[f"D{row}"]
    print(f"Row {row}: B={b.value!r} C={c.value!r} D={d.value!r}")

# Also check B179 formula and related refs on queue sheet
for name in wb.sheetnames:
    if "ПРОФ" in name and "Очередь 1" in name:
        ws = wb[name]
        print("\n--- Cross refs on queue sheet ---")
        for cell in ["B173", "B179", "C33", "E55"]:
            print(f"{cell}={ws[cell].value!r}")
