# -*- coding: utf-8 -*-
import json
import openpyxl

wb = openpyxl.load_workbook(
    'Копия ФС Lite и ПК для предварительной оценки заказного проекта.xlsx',
    data_only=True,
)
out = {'sheets': wb.sheetnames, 'data': {}}

# Key sheets by index/name patterns
targets = []
for i, n in enumerate(wb.sheetnames):
    if any(x in n for x in ['2.', 'Кейс', 'очеред', 'параметр', 'Параметр']):
        targets.append(n)

# Also sheet index 7 = "Оценка свод КейсПроф + очереди" likely
if len(wb.sheetnames) > 7:
    targets.append(wb.sheetnames[7])
if len(wb.sheetnames) > 3:
    targets.append(wb.sheetnames[3])  # параметры расчета

targets = list(dict.fromkeys(targets))

def row_dump(ws, r1, r2, c1=1, c2=15):
    rows = []
    for r in range(r1, r2 + 1):
        cells = {}
        for c in range(c1, c2 + 1):
            v = ws.cell(r, c).value
            if v is not None and v != '':
                cells[openpyxl.utils.get_column_letter(c)] = v
        if cells:
            rows.append({'row': r, 'cells': cells})
    return rows

for name in targets:
    ws = wb[name]
    info = {'name': name, 'rows_1_25': row_dump(ws, 1, 25), 'rows_25_95': row_dump(ws, 25, 95)}
    out['data'][name] = info

with open('scripts/excel_key_sheets.json', 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2, default=str)
print('written scripts/excel_key_sheets.json, sheets:', targets)
