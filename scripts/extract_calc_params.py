# -*- coding: utf-8 -*-
"""Extract calc params and phase details from Excel."""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "Копия ФС Lite и ПК для предварительной оценки заказного проекта.xlsx"

def dump_range(ws, ws_v, r1, r2, c1=1, c2=20):
    lines = []
    for r in range(r1, r2 + 1):
        parts = []
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            val = cell.value
            comp = ws_v.cell(r, c).value
            if val is None and comp is None:
                continue
            addr = f"{get_column_letter(c)}{r}"
            if isinstance(val, str) and val.startswith("="):
                parts.append(f"{addr}={val} => {comp}")
            else:
                disp = val if val is not None else comp
                if disp is not None and str(disp).strip():
                    parts.append(f"{addr}={disp}")
        if parts:
            lines.append(f"R{r}: " + " | ".join(parts))
    return lines

def main():
    wb = load_workbook(XLSX, data_only=False)
    wb_v = load_workbook(XLSX, data_only=True)
    out = []

    # параметры расчета - rates and coefficient tables
    for sn in ["параметры расчета", "Списки"]:
        if sn not in wb.sheetnames:
            continue
        ws, ws_v = wb[sn], wb_v[sn]
        out.append(f"\n# {sn}\n")
        if sn == "параметры расчета":
            out.extend(dump_range(ws, ws_v, 1, 10, 1, 10))
            out.append("\n## Coefficient table ~row 240-250\n")
            out.extend(dump_range(ws, ws_v, 235, 255, 1, 8))
            out.append("\n## Other constants\n")
            out.extend(dump_range(ws, ws_v, 230, 245, 1, 6))
        else:
            out.extend(dump_range(ws, ws_v, 1, 15, 1, 15))
            out.append("\n## K4-K7 project types\n")
            out.extend(dump_range(ws, ws_v, 4, 10, 10, 12))

    for sn in ["Очередь 1 ПРОФ_КОРП", "Очередь 1 Кейс-Совм"]:
        ws, ws_v = wb[sn], wb_v[sn]
        out.append(f"\n# {sn}\n")
        out.append("## Header / FS links rows 1-26\n")
        out.extend(dump_range(ws, ws_v, 1, 26, 1, 15))
        out.append("\n## Team roles row 69-72\n")
        out.extend(dump_range(ws, ws_v, 69, 72, 1, 35))
        out.append("\n## Totals row 91-93\n")
        out.extend(dump_range(ws, ws_v, 90, 93, 1, 30))

    # 1.Содержание проекта - technology criteria
    if "1.Содержание проекта" in wb.sheetnames:
        ws, ws_v = wb["1.Содержание проекта"], wb_v["1.Содержание проекта"]
        out.append("\n# 1.Содержание проекта (technology criteria)\n")
        out.extend(dump_range(ws, ws_v, 1, 30, 1, 8))
        out.extend(dump_range(ws, ws_v, 165, 185, 1, 6))

    # FS sheet org volume
    ws, ws_v = wb["2.ФС для Заполнения"], wb_v["2.ФС для Заполнения"]
    out.append("\n# 2.ФС для Заполнения - org volume AA-AH rows 1-8\n")
    out.extend(dump_range(ws, ws_v, 1, 8, 27, 36))

    Path(ROOT / "scripts/calc_params_detail.txt").write_text("\n".join(out), encoding="utf-8")
    print("written", len(out), "sections")

if __name__ == "__main__":
    main()
