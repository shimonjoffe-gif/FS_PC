# -*- coding: utf-8 -*-
import json
import openpyxl

wb = openpyxl.load_workbook(
    'Копия ФС Lite и ПК для предварительной оценки заказного проекта.xlsx',
    data_only=False,
)

# Find queue sheet
sheet_name = next(n for n in wb.sheetnames if 'ПРОФ_КОРП' in n and 'Очередь 1' in n)
ws = wb[sheet_name]

rows_out = []
for row in range(1, 35):
    vals = {}
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        v = cell.value
        if v is not None:
            from openpyxl.utils import get_column_letter
            vals[get_column_letter(col)] = v
    if vals:
        rows_out.append({'row': row, **vals})

with open('scripts/queue_criteria_extract.json', 'w', encoding='utf-8') as fp:
    json.dump({'sheet': sheet_name, 'rows': rows_out}, fp, ensure_ascii=False, indent=2)

# Also extract rows around criteria table on Содержание - look for ПРОФ/КОРП in column C
ws2 = wb['1.Содержание проекта']
criteria_rows = []
for row in range(95, 200):
    a = ws2.cell(row=row, column=1).value
    b = ws2.cell(row=row, column=2).value
    c = ws2.cell(row=row, column=3).value
    d = ws2.cell(row=row, column=4).value
    e = ws2.cell(row=row, column=5).value
    f = ws2.cell(row=row, column=6).value
    g = ws2.cell(row=row, column=7).value
    if b or c or d or g:
        criteria_rows.append({'row': row, 'A': a, 'B': b, 'C': c, 'D': d, 'E': e, 'F': f, 'G': g})

with open('scripts/content_criteria_extract.json', 'w', encoding='utf-8') as fp:
    json.dump({'rows': criteria_rows}, fp, ensure_ascii=False, indent=2)

print('done', sheet_name, len(criteria_rows))
