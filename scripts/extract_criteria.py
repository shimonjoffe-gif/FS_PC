# -*- coding: utf-8 -*-
import json
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(
    'Копия ФС Lite и ПК для предварительной оценки заказного проекта.xlsx',
    data_only=False,
)

sheet_name = next(n for n in wb.sheetnames if 'Содержание' in n)
ws = wb[sheet_name]

rows_out = []
for row in range(1, 120):
    a = ws.cell(row=row, column=1).value
    b = ws.cell(row=row, column=2).value
    c = ws.cell(row=row, column=3).value
    d = ws.cell(row=row, column=4).value
    e = ws.cell(row=row, column=5).value
    f = ws.cell(row=row, column=6).value
    g = ws.cell(row=row, column=7).value
    if any(v is not None for v in (a, b, c, d, e, f, g)):
        rows_out.append({
            'row': row,
            'A': a, 'B': b, 'C': c, 'D': d, 'E': e, 'F': f, 'G': g,
        })

with open('scripts/criteria_extract.json', 'w', encoding='utf-8') as fp:
    json.dump({'sheet': sheet_name, 'rows': rows_out}, fp, ensure_ascii=False, indent=2)

print('written', len(rows_out), 'rows from', sheet_name)
