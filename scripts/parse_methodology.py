# -*- coding: utf-8 -*-
"""Parse Excel methodology for inventory."""
import json
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "Копия ФС Lite и ПК для предварительной оценки заказного проекта.xlsx"

def cell_info(ws, row, col):
    c = ws.cell(row=row, column=col)
    v = c.value
    f = c.value if (getattr(c, 'data_type', None) == 'f') else None
    if c.data_type == 'f':
        f = str(c.value) if c.value and str(c.value).startswith('=') else None
    # openpyxl stores formula in c.value when data_only=False
    formula = None
    if isinstance(v, str) and v.startswith('='):
        formula = v
        v = None  # will get from data_only sheet later
    return {"row": row, "col": col, "addr": f"{get_column_letter(col)}{row}", "value": v, "formula": formula}

def dump_rows(ws, r1, r2, c1=1, c2=30):
    rows = []
    for r in range(r1, r2 + 1):
        row_data = {"row": r, "cells": []}
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if val is None:
                continue
            entry = {"col": c, "addr": f"{get_column_letter(c)}{r}", "value": val}
            if isinstance(val, str) and val.startswith('='):
                entry["formula"] = val
            rows.append(entry) if False else None
            row_data["cells"].append(entry)
        if row_data["cells"]:
            rows.append(row_data)
    return rows

def main():
    wb = load_workbook(XLSX, data_only=False)
    wb_vals = load_workbook(XLSX, data_only=True)
    print("SHEETS:", wb.sheetnames)

    result = {"sheets": wb.sheetnames, "sheet_analysis": {}}

    # Find sheets with parameter-like content around rows 27-68
    for name in wb.sheetnames:
        ws = wb[name]
        ws_v = wb_vals[name]
        # scan row 27 col A for labels
        labels_27_68 = []
        for r in range(27, 69):
            a = ws.cell(r, 1).value
            b = ws.cell(r, 2).value
            if a or b:
                row_cells = []
                for c in range(1, 25):
                    cv = ws.cell(r, c).value
                    cv_v = ws_v.cell(r, c).value
                    if cv is None and cv_v is None:
                        continue
                    item = {"col": c, "addr": f"{get_column_letter(c)}{r}"}
                    if isinstance(cv, str) and cv.startswith('='):
                        item["formula"] = cv
                        item["computed"] = cv_v
                    else:
                        item["value"] = cv if cv is not None else cv_v
                        if cv_v is not None and cv_v != cv:
                            item["computed"] = cv_v
                    row_cells.append(item)
                if row_cells:
                    labels_27_68.append({"row": r, "cells": row_cells})

        phases_73_90 = []
        for r in range(73, 91):
            row_cells = []
            for c in range(1, 40):
                cv = ws.cell(r, c).value
                cv_v = ws_v.cell(r, c).value
                if cv is None and cv_v is None:
                    continue
                item = {"col": c, "addr": f"{get_column_letter(c)}{r}"}
                if isinstance(cv, str) and cv.startswith('='):
                    item["formula"] = cv
                    item["computed"] = cv_v
                else:
                    item["value"] = cv if cv is not None else cv_v
                    if cv_v is not None and cv_v != cv:
                        item["computed"] = cv_v
                row_cells.append(item)
            if row_cells:
                phases_73_90.append({"row": r, "cells": row_cells})

        if labels_27_68 or phases_73_90:
            result["sheet_analysis"][name] = {
                "params_rows_27_68": labels_27_68,
                "phases_rows_73_90": phases_73_90,
            }

    # Also dump full header area rows 1-26 for context on main calc sheets
    for key_name in ['Оценка свод КейсПроф + очереди', 'параметры расчета', 'Параметры расчета', '2.ФС для Заполнения']:
        if key_name in wb.sheetnames:
            ws = wb[key_name]
            ws_v = wb_vals[key_name]
            header = []
            for r in range(1, 26):
                row_cells = []
                for c in range(1, 35):
                    cv = ws.cell(r, c).value
                    cv_v = ws_v.cell(r, c).value
                    if cv is None and cv_v is None:
                        continue
                    item = {"col": c, "addr": f"{get_column_letter(c)}{r}"}
                    if isinstance(cv, str) and cv.startswith('='):
                        item["formula"] = cv
                        item["computed"] = cv_v
                    else:
                        item["value"] = cv if cv is not None else cv_v
                    row_cells.append(item)
                if row_cells:
                    header.append({"row": r, "cells": row_cells})
            if key_name not in result["sheet_analysis"]:
                result["sheet_analysis"][key_name] = {}
            result["sheet_analysis"][key_name]["header_rows_1_25"] = header

    out = ROOT / "scripts" / "methodology_dump.json"
    out.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding='utf-8')
    print(f"Wrote {out}")
    # print summary
    for sn, data in result["sheet_analysis"].items():
        pr = len(data.get("params_rows_27_68", []))
        ph = len(data.get("phases_rows_73_90", []))
        if pr or ph:
            print(f"  {sn}: params={pr} phase_rows={ph}")

if __name__ == "__main__":
    main()
