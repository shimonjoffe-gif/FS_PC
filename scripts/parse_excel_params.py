# -*- coding: utf-8 -*-
import json
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(
    'Копия ФС Lite и ПК для предварительной оценки заказного проекта.xlsx',
    data_only=False,
)
print('SHEETS:', wb.sheetnames)

# Find sheet with params around row 27
for name in wb.sheetnames:
    ws = wb[name]
    a27 = ws.cell(27, 1).value
    if a27 and ('параметр' in str(a27).lower() or 'технолог' in str(a27).lower() or 'ставка' in str(a27).lower()):
        print(f'\n=== CANDIDATE: {name} A27={a27!r} ===')

# Try sheets that look like main eval sheet
target_names = [n for n in wb.sheetnames if 'оценк' in n.lower() or 'предвар' in n.lower() or n.startswith('1.')]
if not target_names:
    target_names = wb.sheetnames[2:6]

for name in target_names[:5]:
    ws = wb[name]
    print(f'\n========== SHEET: {name} ==========')
    for r in range(1, 100):
        cells = []
        for c in range(1, 12):
            cell = ws.cell(r, c)
            v = cell.value
            if v is not None:
                f = ''
                if isinstance(v, str) and v.startswith('='):
                    f = ' [F]'
                cells.append(f'{get_column_letter(c)}{r}={v!r}{f}')
        if cells:
            print(f'R{r}:', ' | '.join(cells))
